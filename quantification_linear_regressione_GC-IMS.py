# 📦 Librerie
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from google.colab import files

# 📥 Upload file Excel
uploaded = files.upload()
file_path = next(iter(uploaded))
df = pd.read_excel(file_path)

# 🔍 Separa calibrazione da campioni
split_index = df[df.iloc[:, 0].astype(str).str.upper() == 'SAMPLES'].index[0]
calibration_df = df.iloc[:split_index].reset_index(drop=True)
sample_df = df.iloc[split_index + 1:].reset_index(drop=True)

# 🎯 Target e analiti
y_calib = calibration_df['concentration'].astype(float).values
analiti = sorted(set(col.split('_')[0] for col in calibration_df.columns if '_H_m' in col))
colors = plt.cm.get_cmap("tab10", len(sample_df))

# 📊 Contenitori risultati
sample_predictions = []
coefficients_data = []
plot_images = {}

# 🔁 Loop analiti
for analita in analiti:
    try:
        # X Calibrazione
        X_calib = calibration_df[[f"{analita}_H_m", f"{analita}_V_m", f"{analita}_H_d", f"{analita}_V_d"]].astype(float).values
        model = LinearRegression()
        model.fit(X_calib, y_calib)
        y_pred_calib = model.predict(X_calib)
        r2 = r2_score(y_calib, y_pred_calib)

        # Salva coeff.
        coefficients_data.append({
            "Analita": analita,
            "Coeff_H_m": model.coef_[0],
            "Coeff_V_m": model.coef_[1],
            "Coeff_H_d": model.coef_[2],
            "Coeff_V_d": model.coef_[3],
            "Intercept": model.intercept_,
            "R2": r2
        })

        # 🔄 Grafico calibrazione + campioni + legenda separata
        fig, (ax, ax_legend) = plt.subplots(1, 2, figsize=(10, 5), gridspec_kw={"width_ratios": [3, 1]})
        
        ax.scatter(y_calib, y_pred_calib, label="Calibrazione", color='black')
        ax.plot([min(y_calib), max(y_calib)], [min(y_calib), max(y_calib)], 'k--')

        predicted_vals = {}
        legend_elements = []
        for idx, row in sample_df.iterrows():
            sample_name = row.iloc[0]
            color = colors(idx)
            X_sample = row[[f"{analita}_H_m", f"{analita}_V_m", f"{analita}_H_d", f"{analita}_V_d"]].astype(float).values.reshape(1, -1)
            y_sample_pred = model.predict(X_sample)[0]
            y_for_plot = max(y_sample_pred, 0)
            label = "<LQ" if y_sample_pred < 0 else f"{y_sample_pred:.3f}"
            predicted_vals[sample_name] = label

            ax.scatter([0], [y_for_plot], color=color)
            ax.plot([0, 0], [0, y_for_plot], linestyle='dotted', color=color)
            ax.plot([0, y_for_plot], [y_for_plot, y_for_plot], linestyle='dotted', color=color)

            legend_elements.append(plt.Line2D([0], [0], marker='o', color='w', label=sample_name,
                                              markerfacecolor=color, markersize=8))

        ax.set_xlabel("Concentrazione reale")
        ax.set_ylabel("Concentrazione predetta")
        ax.set_title(f"{analita} - R² = {r2:.3f}")
        ax.set_aspect('equal', adjustable='box')
        ax.grid(True)

        ax_legend.axis('off')
        ax_legend.legend(handles=legend_elements, loc='upper left', frameon=True)
        fig.tight_layout()

        # Salva immagine
        buf = BytesIO()
        fig.savefig(buf, format='png')
        buf.seek(0)
        plot_images[analita] = buf
        plt.close(fig)

        # Salva previsioni
        for k, v in predicted_vals.items():
            existing = next((item for item in sample_predictions if item['sample'] == k), None)
            if existing:
                existing[analita] = v
            else:
                sample_predictions.append({'sample': k, analita: v})

    except Exception as e:
        print(f"⚠️ Errore per {analita}: {e}")

# 🧾 Costruisci DataFrame risultati
results_df = pd.DataFrame(sample_predictions)
coefficients_df = pd.DataFrame(coefficients_data)

# 📝 Crea file Excel
output_path = "RESULTS_" + file_path
wb = Workbook()

# Sheet 1: results
ws_results = wb.active
ws_results.title = "results"
for r in dataframe_to_rows(results_df, index=False, header=True):
    ws_results.append(r)

# Sheet 2: original data
ws_data = wb.create_sheet("original_data")
for r in dataframe_to_rows(df, index=False, header=True):
    ws_data.append(r)

# Sheet 3: coefficients
ws_coeff = wb.create_sheet("linear equations coefficients")
for r in dataframe_to_rows(coefficients_df, index=False, header=True):
    ws_coeff.append(r)

# Sheet 4: calibration plots
ws_plots = wb.create_sheet("calibration plots")
row_pos = 1
for analita, image_buf in plot_images.items():
    img = XLImage(PILImage.open(image_buf))
    img.width = 600
    img.height = 600
    ws_plots.add_image(img, f"A{row_pos}")
    row_pos += 35

# 💾 Salva e scarica
wb.save(output_path)
files.download(output_path)
